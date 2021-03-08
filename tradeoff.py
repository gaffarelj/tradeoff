# Author: Jeremie Gaffarel, lastly revised on the 4th of March 2021
# This module contains the class to run the tradeoff, the seneitivity
# analysis, and save the results
import numpy as np
import pandas as pd
import warnings
import copy
from matplotlib import pyplot as plt
warnings.filterwarnings('ignore')

class color:
	"""
	Save the color code and name of any color
	"""
	def __init__(self, code, name):
		self.HTML = code
		self.name = name	

	def __repr__(self):
		# When using str(color) or print(color), only return its name
		return self.name

def val_s(number):
	"""
	Convert a number to string
	"""
	# If the number is a nan (or an empty string), return an empty symbol
	if type(number) == np.str_:
		if number == "":
			return "∅"
		return number
	elif np.isnan(number):
		return "∅"
	# Start the resulting string by a dash if the number is negative
	res = "" if number >= 0 else "-"
	# Take the absolue value of the number
	number = np.fabs(number)
	# If the number is too small, return 0
	if number < 10 ** (-22):
		res += str(0)
	# If the number is between 1000 and 1/1000, return it rounder to three decimals
	elif number < 1e3 and number > 0.001:
		res+= str(round(number, 3))
	# Else use a scientific notation
	else:
		res += "{:.2e}".format(number)
	# Remove everything after the decimal if it's just .0
	if res[-2:] == ".0":
		res = res[:-2]
	return res

class tradeoff:
	"""
	Tradeoff class, to normalise the data bewteen the best and worst, run the 
	tradeoff, optionnaly run a sensitivity analysis, and save the output
	Inputs:
	 * data (pandas dataframe): data on which to run the tradeoff
	 * config (list[open_file.d_conf]): configuration of each of the columns
	 * worst_best (dict[tuple]): dictionnary of the worst and best values of each column
	 * colors (list[tradeoff.color]): list of colors for the output table
	"""
	def __init__(self, data, config, worst_best, colors=None):
		# Save the data before any modification is done
		self.rawdata = data
		# Copy the data to make modifications
		self.data = data.copy()
		# Save the columns config
		self.config = config
		# Get the name of all columns
		self.names = self.get_names()
		# Save the worst and best of all columns
		self.worst_best = worst_best
		# Declare the sensitivity dictionnary as a None
		self.sens_dict = None
		# If the colors are not specified, set them as default ones
		if colors is None:
			self.colors = [color("EF5350", "red"), color("FB8C00", "orange"), color("FFEB3B", "yellow"), color("8BC34A", "green"), color("00BCD4", "blue")]
		else:
			self.colors = colors
		# Get the mean and standard deviation for each column
		self.get_stats()
		# Go trough all columns configurations
		for i, col_conf in enumerate(self.config):
			# If the column is not to be included in the tradeoff...
			if not col_conf.in_to():
				# ... remove it from the dataset
				self.data.drop(columns=col_conf.name, inplace=True)
		# Normalise the weights
		self.norm_weights()
		# Normalise the data of all columns
		self.normalise_all()
		# Run the tradeoff
		self.run_tradeoff()

	def norm_weights(self):
		"""
		This function ensures that the sum of all weights is equal to 1 (100%)
		"""
		# Get all weights that are from columns included in the tradeoff in a list
		valid_weights = [conf.weight if conf.in_to() else 0 for conf in self.config]
		# Devide all column/parameters weights by the sum of the weights
		for para in self.config:
			if para.in_to():
				para.weight /= sum(valid_weights)

	def normalise(self, val, param):
		"""
		Normalise the data between the best and worst value of a column
		Inputs:
		 * val (float/bool): value to normalise
		 * param (open_file.d_col): column parameters
		"""
		# Extract the parameter (column) name
		pname = param.name
		# If the value is a nan (no value was given in the Excel sheet)...
		is_nan = False
		if type(val) == np.str_:
			if val == "":
				is_nan = True
		elif np.isnan(val):
			is_nan = True
		if is_nan:
			# ... and if the mean is nan, then the column is empty, return 0.5
			if np.isnan(param.mean):
				return 0.5
			# ... set the value as the column mean
			val = param.mean
		# If the column contains a range, evaluate the normalised value based on it
		if param.range is not None:
			# Split the value at the dash, in case two values are present
			vals = [float(v) for v in str(val).split("-")]
			# Take 100 samples to make sure to catch if any value is in the range
			for _val in np.linspace(min(vals), max(vals), 100):
				# Return 1 if the value is within the range
				if float(_val) <= max(param.range) and float(_val) >= min(param.range):
					return 1
			# Otherwise, return 0
			return 0
		# Get the worst and best value of the column
		w_b = self.worst_best[pname]
		# If the worst/best is a bool...
		if type(w_b) == bool:
			# ... return 1 if the value equals the worst/best
			if val == w_b:
				return 1
			# ... return 0 otherwise
			return 0
		# Separate the worst and best from the tuple
		worst, best = w_b
		# If the worst equals the best, return 1
		if worst == best:
			return 0.5
		# Linearly normalise the value between the worst and best values
		norm = (val - worst) / (best - worst)
		# Force the normalised value to be between 0 and 1
		if norm <= 0:
			norm = 0
		elif norm >= 1:
			norm = 1
		return norm

	def normalise_all(self):
		"""
		Normalise all od the values from the datafrome
		"""
		# Go trough each parameter
		for param in self.config:
			if param.in_to():
				# Extract the values from the dataframe as an array,
				# and save as a list in all parameters
				param.val_in = np.array([row[param.name] for _, row in self.data.iterrows()])
				# Normalise all values, and also save in each parameter as a list
				param.val_norm = np.array([self.normalise(val, param) for val in param.val_in])

	def run_tradeoff(self):
		"""
		Run the tradeoff in itself
		"""
		# Prepare an empty array to contain the tradeoff results
		self.total = np.zeros(len(self.data))
		# Go trough each column
		for i, param in enumerate(self.config):
			# Check if the column is to be included in the tradeoff
			if param.in_to():
				# Add the column values multiplied by the weights to the tradeoff results
				self.total += param.val_norm*param.weight

	def get_names(self):
		"""
		Get the custom name of each row from the data
		"""
		names_list = []
		# Go trough each column
		for param in self.config:
			# If the weight is specified, but is negative,
			# this indicates that we want this to be in the name
			if param.weight is not None and param.weight < 0:
				# Gather the data of all rows for the given column
				names_list.append(self.data[param.name].astype(str).values.tolist())
		# Transpose the nested list
		names_list = [list(i) for i in zip(*names_list)]
		# Join the different part of the names with : in between
		names = [":".join(names) for names in names_list]
		return names

	def get_stats(self):
		"""
		Get the mean and standard deviation for each column
		"""
		# Go trough each column
		for param in self.config:
			# If the column is included in the tradeoff
			if param.in_to():
				# If the column contains booleans, set the mean and standard deviation as nan
				if type(param.best) == bool:
					param.std, param.mean = np.nan, np.nan
				# Otherwise, compute it from the column data
				else:
					# If the column contain ranges, unpack them before computing the mean and std
					if param.range is not None:
						vals = [str(v).split("-") for v in self.data[param.name]]
						vals = [float(_v) for v in vals for _v in v]
						param.std, param.mean = np.nanstd(vals), np.nanmean(vals)
					else:
						param.std, param.mean = self.data[param.name].std(), self.data[param.name].mean()
				# If the standard deviation is 0, all values are identical
				if param.std == 0:
					# Set the weight as None to exclude the column from the tradeoff
					param.weight = None

	def run_sensitivity(self, n, std_frac, box_plot=False):
		"""
		Run the sensitivity analysis
		Inputs:
		 * n (int): number of tradeoffs to run
		 * stf_frac (float): standard deviation = stf_frac * mean
		"""
		# Make sure the number of tradeoff to run is an int
		n = int(n)
		# Get the weights of all columns/params
		weights = [param.weight for param in self.config if param.in_to()]
		# Create a dict to save the winner of each tradeoff
		best_dict = dict()
		best_weights_dict = dict()
		max_weight = 0
		print("Running weight sensitivity analysis...")
		# Repeat the following n times...
		for i in range(n):
			# Every 100 runs, print a progress indicator
			if (i + 1) % 100 == 0:
				print("%s/%s (%.1f%%) done..." % (i+1, n, (i+1)/n*100), end="\r")
			# (Deep-)copy the current tradeoff
			to_copy = copy.deepcopy(self)
			# Go trough each column
			for j, param in enumerate(to_copy.config):
				if param.in_to():
					# Get the set weight
					old_weight = self.config[j].weight
					# Change the weight to one randomly picked from a normal distribution
					# centred at the old weight, with a standard deviation of stf_frac * the old weight
					new_weight = max(np.random.normal(old_weight, old_weight*std_frac), 0)
					param.weight = new_weight
			# Normalise all weights
			to_copy.norm_weights()
			# Run the tradeoff with the new weights
			to_copy.run_tradeoff()
			# Get the winner from this tradeoff
			best_id = to_copy.get_output(format="return best")
			# Increment the save dict by 1/n for the current winner
			if best_id in best_dict.keys():
				best_dict[best_id] += 1/n
			else:
				best_dict[best_id] = 1/n
			# Save the weights of this tradeoff that made this specific concept win
			valid_weights = [conf.weight for conf in to_copy.config if conf.in_to()]
			max_weight = max(max_weight, max(valid_weights))
			if best_id in best_weights_dict.keys():
				best_weights_dict[best_id].append([w*100 for w in valid_weights])
			else:
				best_weights_dict[best_id] = [[w*100 for w in valid_weights]]
			# Delete the tradeoff copy (save some memory)
			del to_copy
		print()
		# Save the column name for the excel output sheet
		self.sens_col_name = "Sensitivity \n %.1e runs \n σ = %.2f μ" % (n, std_frac)
		# Save the sensitivity dictionnary, containing the percentage
		# of times each concept won the tradeoff
		self.sens_dict = best_dict
		# Plot a boxplot of the weights that make each tradeoff win
		if box_plot:
			print("Plotting weights that win the sensitivity analysis...")
			# Go trough each concept that won the tradeoff
			for key, val in best_weights_dict.items():
				# Extract the column names
				col_n = [conf.name for conf in self.config if conf.in_to()]
				# Make a boxplot of the winning weights for this concept
				plt.boxplot(np.array(val), showfliers=False)
				plt.xticks(np.arange(1, len(col_n)+1, 1), col_n, rotation="vertical")
				plt.title("%s (wins %s%%)" % (key, val_s(best_dict[key]*100))), plt.ylabel("Winning weights [%]")
				# Also plot the weight used out of the sensitivity analysis
				for i, conf in enumerate([c for c in self.config if c.in_to()]):
					plt.plot(i+1, conf.weight*100, marker="_", c="green", markersize=17)
				plt.plot([], [], marker="_", c="green", label="Nominal weight")
				plt.plot([], [], marker="_", c="orange", label="Sensitivity median")
				plt.tight_layout(), plt.grid(), plt.ylim(0, int(max_weight*100)+1)
				plt.legend(loc="upper right"), plt.savefig("plots/%s.pdf" % key, dpi=300), plt.close()
		
	def get_output(self, format="python", fname=None):
		"""
		Get the output of the tradeoff
		Inputs:
		 * format (["python", "return best", "excel"]): choose in which format to return the results
		 * fname (string): in which Excel file to save the tradeoff table (only necessary for "excel")
		"""
		# If the format is "return best"
		if format == "return best":
			# Get the index of the best scoring row
			idx_best = np.where(self.total == max(self.total))[0][0]
			# Return the best scoring row id
			return self.rawdata["ID"][idx_best]
		# If the format is "python"
		elif format == "python":
			# Go trough each column
			for param in self.config:
				if param.in_to():
					# Print the parameter value and normalised value for each row
					print(param.name, ", \t actual value:", "\t".join([str(round(val, 3)) for val in param.val_in]), "\n")
					print(param.name, ", \t scaled value:", "\t".join([str(round(val, 3)) for val in param.val_norm]), "\n")
			# Print the tradeoff scores
			print(" * final values:", "\t".join([str(round(val, 3)) for val in self.total]), "\n")
			# If the sensitivity dictionnary is not None, print it
			if self.sens_dict is not None:
				print(" * sensitivity analysis:", self.sens_dict)
			print()
		# If the format is "excel"
		elif format == "excel":
			# Make a copy of the data
			out_table = self.data.copy()
			# Create a dictionnary to contain the new table header
			new_header = dict()
			# Go trough each column
			for param in self.config:
				if param.in_to():
					# Set the parameter colors
					param.set_colors(self.colors)
					# Get the worst/best value of the column
					w_b = self.worst_best[param.name]
					# If the worst/best is a boolean
					if type(w_b) == bool:
						# Set the worst/best message as "True/False is best"
						w_b_s = "%s is best" % w_b
						worst_best = "True, False"
						head = "%s \n %.2f%% \n (%s) \n %s" % (param.name, param.weight*100, \
							worst_best, w_b_s)
					# If a range is specified, print that it's best to be in it
					elif param.range is not None:
						w_b_s = "Best in \n [%s, %s]" % ( val_s(min(param.range)), val_s(max(param.range)) )
						head = "%s \n %.2f%% \n %s \n μ = %s \n σ = %s " % (param.name, param.weight*100, \
							w_b_s, val_s(param.mean), val_s(param.std))
					# If worst/best is a min/max, set the message as "High/Low best"
					else:
						if param.best == max:
							w_b_s = "High best"
						else:
							w_b_s = "Low best"
						worst_best = "%s, %s" % (val_s(min(w_b)), val_s(max(w_b)))
						head = "%s \n %.2f%% \n (%s) \n %s \n μ = %s \n σ = %s " % (param.name, param.weight*100, \
							worst_best, w_b_s, val_s(param.mean), val_s(param.std))
					# Save the column header
					new_header[param.name] = head
					# Set each of the table cells as "value -> normalised value" instead of just "value"
					for i in range(len(param.val_in)):
						s = val_s(param.val_in[i]) + " → " + val_s(round(param.val_norm[i], 3))
						out_table[param.name][i] = s
			# Replace the old headers by the new ones
			out_table = out_table.rename(columns=new_header)
			# Insert an updated ID column at the beginning of the table
			out_table.insert(0, "ID", self.names, True) 
			# Insert a column with the scores at the end of the table
			out_table.insert(len(out_table.columns), "Score", [round(score, 5) for score in self.total], True) 
			# If the sensibility dict is not None
			if self.sens_dict is not None:
				# Set the whole sensitivity row as 0%
				sens_col = [str(round(0.0, 3)) + "%"] * len(self.total)
				# Go trough each value of the sensitivity dict
				for key, value in self.sens_dict.items():
					# Get the row index of the given row name (key) from the dict
					idx_val = np.where(self.rawdata["ID"] == key)[0][0]
					# Set the row value as the percentage from the dict
					sens_col[idx_val] = str(round(value * 100, 3)) + "%"
				# Insert the sensitivity column at the end of the table
				out_table.insert(len(out_table.columns), self.sens_col_name, sens_col, True)
			# Get the best score from the tradeoff
			best_score = max(out_table["Score"])
			# Get the concept associated with the best score
			best_concept = out_table.loc[out_table['Score'] == best_score].iloc[0, 0]
			# Print the best concept and its score
			print("Best concept according to the trade-off is %s with a score of %s" % (best_concept, best_score))
			# Import modules to make the table much nicer
			from openpyxl import load_workbook
			from openpyxl.styles import Alignment, Font, PatternFill
			from openpyxl.utils import get_column_letter
			from openpyxl.utils.dataframe import dataframe_to_rows
			from openpyxl.worksheet.dimensions import ColumnDimension
			from openpyxl.comments import Comment
			# Open the excel file containing the data
			wb=load_workbook(fname)
			# If the result sheet already exists, remove it
			try:
				ws = wb["Trade-off results"]
				wb.remove_sheet(ws)
			except KeyError:
				pass
			# Else, create it
			ws = wb.create_sheet("Trade-off results")
			# Add all rows from the result table to the worksheet
			for r in dataframe_to_rows(out_table, index=False, header=True):
				ws.append(r)
			# Go trough each cell of each column
			i_col = 0
			for param in self.config:
				if param.in_to():
					i_col += 1
					for i, row in enumerate(ws.iter_rows()):
						if i > 0:
							# And set the background color as the one that was selected according
							# to the normalised value
							cell = row[i_col]
							cell.fill = PatternFill(fill_type='solid', start_color=param.color[i-1].HTML)
			# Create a dict to save the column widths
			dims = dict()
			# Go trough each cell of each row
			for i, row in enumerate(ws.rows):
				for cell in row:
					# Center the text
					cell.alignment = Alignment(wrapText=True, horizontal="center", vertical="center")
					# If the row is the header one, make the text bold
					if i == 0:
						cell.font = Font(bold=True)
					max_val = 0
					# If the cell contains something, and is not in the header
					if cell.value and i > 0:
						# Count the number of letters in the cell, save it if its the max in the row
						max_val = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
					else:
						# For the header, do the same but doing each line seperately, splitting the line breaks
						for line in cell.value.split(" \n "):
							max_val = max(len(line), max_val, 7)
					# Save the best column width
					dims[cell.column_letter] = max_val
			# Go trough each column and width from the dict
			for col, value in dims.items():
				# Apply the width to each column
				ws.column_dimensions[col].width = value
			# Set the sheet zoom to 90% (makes it fits completely on the screen; to be tuned if needed)
			ws.sheet_view.zoomScale = 90
			# Save the sheet
			wb.save(fname)
		# If the format was not in ["python", "return best", "excel"], print a warning
		else:
			print("Warning: output format is unknown.")